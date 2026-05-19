"""
excel_report.py
---------------
Generates a professional multi-sheet Excel workbook:
  Sheet 1 – Raw Data
  Sheet 2 – Monthly Summary
  Sheet 3 – Category Analysis
  Sheet 4 – Budget vs Actual
  Sheet 5 – Scenario Analysis

Run standalone:
    python scripts/excel_report.py
"""

import os, sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

sys.path.insert(0, os.path.dirname(__file__))
from data_cleaning import load_and_clean, CLEAN_PATH
from analysis import monthly_spending_trend, category_analysis, budget_vs_actual, scenario_analysis

OUT_PATH = os.path.join(os.path.dirname(__file__), "..", "outputs", "excel",
                        "Financial_Analysis_Report.xlsx")
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

# ── Style helpers ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1F4E79")   # dark blue
SUBHEAD_FILL  = PatternFill("solid", start_color="2E75B6")   # medium blue
ALT_FILL      = PatternFill("solid", start_color="D6E4F0")   # light blue alt row
GREEN_FILL    = PatternFill("solid", start_color="C6EFCE")
RED_FILL      = PatternFill("solid", start_color="FFC7CE")
YELLOW_FILL   = PatternFill("solid", start_color="FFEB9C")

WHITE_BOLD    = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
DARK_BOLD     = Font(bold=True, color="1F4E79", name="Calibri", size=11)
REGULAR       = Font(name="Calibri", size=10)
CURRENCY_FMT  = '#,##0.00'
PCT_FMT       = '0.00%'

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def _header(ws, row, col, value, width=18):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = WHITE_BOLD
    c.fill   = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _cell(ws, row, col, value, fmt=None, fill=None, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = Font(name="Calibri", size=10, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    return c


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _build_raw(wb, df):
    ws = wb.create_sheet("Raw Data")
    cols = ["Date","Month","Expense_Category","Amount","Income","Budget",
            "Payment_Method","Year","Quarter","Budget_Variance","Savings_Rate"]
    cols = [c for c in cols if c in df.columns]

    for ci, col in enumerate(cols, 1):
        _header(ws, 1, ci, col)

    for ri, (_, row) in enumerate(df[cols].iterrows(), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, col in enumerate(cols, 1):
            val = row[col]
            fmt = None
            if col in ("Amount","Income","Budget","Budget_Variance"):
                fmt = CURRENCY_FMT
            elif col == "Savings_Rate":
                fmt = '0.00'
            _cell(ws, ri, ci, val, fmt=fmt, fill=fill)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df)+1}"


def _build_monthly(wb, df):
    ws = wb.create_sheet("Monthly Summary")
    monthly = monthly_spending_trend(df).copy()
    monthly["Month"] = monthly["Month"].dt.strftime("%Y-%m")

    headers = ["Month","Income","Total_Expense","Savings","Savings_Rate (%)"]
    for ci, h in enumerate(headers, 1):
        _header(ws, 1, ci, h, width=16)

    for ri, (_, row) in enumerate(monthly.iterrows(), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        _cell(ws, ri, 1, row["Month"],         fill=fill, align="center")
        _cell(ws, ri, 2, row["Income"],         fmt=CURRENCY_FMT, fill=fill, align="right")
        _cell(ws, ri, 3, row["Total_Expense"],  fmt=CURRENCY_FMT, fill=fill, align="right")

        sav_fill = GREEN_FILL if row["Savings"] >= 0 else RED_FILL
        _cell(ws, ri, 4, row["Savings"],        fmt=CURRENCY_FMT, fill=sav_fill, align="right")

        rate_fill = GREEN_FILL if row["Savings_Rate"] >= 10 else (RED_FILL if row["Savings_Rate"] < 0 else None)
        _cell(ws, ri, 5, row["Savings_Rate"],   fmt='0.00', fill=rate_fill, align="right")

    # Totals row
    n = len(monthly) + 1
    tr = n + 1
    _cell(ws, tr, 1, "TOTAL / AVG", bold=True, fill=SUBHEAD_FILL, align="center")
    ws.cell(row=tr, column=2, value=f'=SUM(B2:B{n})').number_format = CURRENCY_FMT
    ws.cell(row=tr, column=3, value=f'=SUM(C2:C{n})').number_format = CURRENCY_FMT
    ws.cell(row=tr, column=4, value=f'=SUM(D2:D{n})').number_format = CURRENCY_FMT
    ws.cell(row=tr, column=5, value=f'=AVERAGE(E2:E{n})').number_format = '0.00'
    for ci in range(1, 6):
        ws.cell(row=tr, column=ci).font  = DARK_BOLD
        ws.cell(row=tr, column=ci).border = THIN_BORDER

    # Line chart: Income vs Expense
    chart = LineChart()
    chart.title = "Monthly Income vs Expense"
    chart.style = 10
    chart.height = 12
    chart.width  = 24
    income_ref  = Reference(ws, min_col=2, min_row=1, max_row=n)
    expense_ref = Reference(ws, min_col=3, min_row=2, max_row=n)
    chart.add_data(income_ref,  titles_from_data=True)
    chart.add_data(expense_ref, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.solidFill = "2ECC71"
    chart.series[1].graphicalProperties.line.solidFill = "E74C3C"
    ws.add_chart(chart, f"G2")
    ws.freeze_panes = "A2"


def _build_category(wb, df):
    ws = wb.create_sheet("Category Analysis")
    cat = category_analysis(df)

    headers = ["Category","Total Spent","Avg Monthly","Budget","Variance","% of Total","Transactions","Status"]
    widths  = [18, 14, 14, 12, 12, 12, 14, 18]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _header(ws, 1, ci, h, width=w)

    for ri, (_, row) in enumerate(cat.iterrows(), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        _cell(ws, ri, 1, row["Expense_Category"],  fill=fill)
        _cell(ws, ri, 2, row["Total_Spent"],        fmt=CURRENCY_FMT, fill=fill, align="right")
        _cell(ws, ri, 3, row["Avg_Monthly"],        fmt=CURRENCY_FMT, fill=fill, align="right")
        _cell(ws, ri, 4, row["Budget"],             fmt=CURRENCY_FMT, fill=fill, align="right")

        vfill = GREEN_FILL if row["Budget_Variance"] >= 0 else RED_FILL
        _cell(ws, ri, 5, row["Budget_Variance"],    fmt=CURRENCY_FMT, fill=vfill, align="right")
        _cell(ws, ri, 6, row["Pct_of_Total"],       fmt='0.00', fill=fill, align="right")
        _cell(ws, ri, 7, row["Transaction_Count"],  fill=fill, align="center")
        sfill = GREEN_FILL if "Under" in str(row["Status"]) else RED_FILL
        _cell(ws, ri, 8, row["Status"],             fill=sfill, align="center")

    ws.freeze_panes = "A2"


def _build_budget_vs_actual(wb, df):
    ws = wb.create_sheet("Budget vs Actual")
    bva = budget_vs_actual(df)

    headers = ["Month","Category","Budget","Actual","Variance","% Used","Over Budget?"]
    widths  = [12, 18, 13, 13, 13, 10, 14]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _header(ws, 1, ci, h, width=w)

    for ri, (_, row) in enumerate(bva.iterrows(), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        _cell(ws, ri, 1, row["Month"],             fill=fill, align="center")
        _cell(ws, ri, 2, row["Expense_Category"],  fill=fill)
        _cell(ws, ri, 3, row["Budget"],            fmt=CURRENCY_FMT, fill=fill, align="right")
        _cell(ws, ri, 4, row["Actual"],            fmt=CURRENCY_FMT, fill=fill, align="right")

        vfill = GREEN_FILL if row["Variance"] >= 0 else RED_FILL
        _cell(ws, ri, 5, row["Variance"],          fmt=CURRENCY_FMT, fill=vfill, align="right")
        pfill = RED_FILL if row["Pct_Used"] > 100 else fill
        _cell(ws, ri, 6, row["Pct_Used"],          fmt='0.00', fill=pfill, align="right")
        obfill = RED_FILL if row["Over_Budget"] else GREEN_FILL
        _cell(ws, ri, 7, "YES" if row["Over_Budget"] else "NO", fill=obfill, align="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(7)}{len(bva)+1}"


def _build_scenario(wb, df):
    ws = wb.create_sheet("Scenario Analysis")

    scenarios = [
        ("Baseline",           0,   0),
        ("+10% Salary",       10,   0),
        ("+20% Salary",       20,   0),
        ("-10% Salary",      -10,   0),
        ("+10% Expenses",      0,  10),
        ("+20% Expenses",      0,  20),
        ("+10% Sal +10% Exp", 10,  10),
        ("+20% Sal -10% Exp", 20, -10),
    ]

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "🔮 Financial Scenario / What-If Analysis"
    ws["A1"].font  = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["Scenario","Salary Δ%","Expense Δ%","New Income","New Expense",
               "New Savings","Savings Rate%","Impact"]
    widths  = [22, 12, 13, 14, 14, 14, 14, 14]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _header(ws, 2, ci, h, width=w)

    for ri, (name, sal_chg, exp_chg) in enumerate(scenarios, 3):
        res = scenario_analysis(df, sal_chg, exp_chg)
        s   = res["scenario"]
        fill = ALT_FILL if ri % 2 == 0 else None
        impact_fill = GREEN_FILL if "Positive" in s["impact"] else RED_FILL

        _cell(ws, ri, 1, name,             fill=fill, bold=(ri==3))
        _cell(ws, ri, 2, sal_chg,          fmt='0.0', fill=fill,        align="center")
        _cell(ws, ri, 3, exp_chg,          fmt='0.0', fill=fill,        align="center")
        _cell(ws, ri, 4, s["new_income"],  fmt=CURRENCY_FMT, fill=fill, align="right")
        _cell(ws, ri, 5, s["new_expense"], fmt=CURRENCY_FMT, fill=fill, align="right")

        sf = GREEN_FILL if s["new_savings"] >= 0 else RED_FILL
        _cell(ws, ri, 6, s["new_savings"],      fmt=CURRENCY_FMT, fill=sf,          align="right")
        rf = GREEN_FILL if s["new_savings_rate"] >= 10 else RED_FILL
        _cell(ws, ri, 7, s["new_savings_rate"], fmt='0.00',        fill=rf,          align="right")
        _cell(ws, ri, 8, s["impact"],           fill=impact_fill,  align="center")

    # Assumptions note
    note_row = 3 + len(scenarios) + 1
    ws.merge_cells(f"A{note_row}:H{note_row}")
    ws[f"A{note_row}"] = ("ℹ  Assumptions: All changes are applied to the 2-year average monthly "
                           "income/expense baseline.  Savings Rate = (New Savings / New Income) × 100.")
    ws[f"A{note_row}"].font = Font(name="Calibri", size=9, italic=True, color="666666")
    ws[f"A{note_row}"].fill = YELLOW_FILL


def build_excel_report(df: pd.DataFrame = None):
    if df is None:
        if os.path.exists(CLEAN_PATH):
            df = pd.read_csv(CLEAN_PATH, parse_dates=["Date"])
        else:
            df = load_and_clean()

    wb = Workbook()
    wb.remove(wb.active)          # remove default empty sheet

    _build_raw(wb, df)
    _build_monthly(wb, df)
    _build_category(wb, df)
    _build_budget_vs_actual(wb, df)
    _build_scenario(wb, df)

    wb.save(OUT_PATH)
    print(f"✅  Excel report saved → {os.path.abspath(OUT_PATH)}")


if __name__ == "__main__":
    build_excel_report()
